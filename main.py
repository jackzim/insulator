from logging import raiseExceptions
import numpy as n
import re
import os
from itertools import chain
import openpyxl as xl
from frontend import *
from functions import *
from validity_check import *

## Trystar Electrical Part Program
## Mantained by JMZ
## Supports:
## -> Adders
## -> Breakers, Lug Kits
## -> Breaker and Transfer Switch Aux Contacts
## -> Phase Rotation Monitors and Fillers (if needed)
## -> Internal Wiring for Compression Lugs
## -> ERMS, 100% Rated
## -> Trystar Manual Transfer Switches
## -> 4P Support for Wall Mounts
## Correct: py -m PyInstaller --onefile --icon=image/trystar-mark.ico --noconsole main.py

## Current Work in Progress:
## -> TATS Support
## -> Multiple Adders
## -> Idea for multiple adder: Replace 2 with previous entry of list

print(one_line)
print(part_number)

i = 0

## Excel Setup

excel_wb = xl.Workbook()
excel_sheet = excel_wb.active

## Calling Functions

if __name__ == "__main__":
    part_number = part_number.upper()
    part_list = strip_number(part_number)
    amperage, voltage, voltage_code, enclosure_code = electrical_coding(part_list)
    purpose_validity = purpose_matching(part_list, one_line)
    modifier_validity = modifier_valid(one_line, erms, is_100_percent_rated)
    checkbox_validity = invalid_checkbox(is_4P, is_100_percent_rated, enclosure_code)
    adders = identify_adders(part_list, voltage_code)
    adder_parts = adder_part_list(adders)
    prm_list = phase_rotation_monitor(voltage_code)
    print(f"Amperage: {amperage}, Voltage: {voltage}")

    if enclosure_code in ["W", "L"]:
        breaker_kit, quantity = wallmount_breaker_pull(one_line, amperage, is_100_percent_rated, is_4P)
        temp_gen_kit = []
    elif enclosure_code == "P":
        temp_gen_kit = temp_gen_hardwire[str(amperage)]
        if part_list[0] in ["TMTS", "TATS"]:
            breaker_kit, quantity = wallmount_breaker_pull(one_line, amperage, is_100_percent_rated)
        elif part_list[0] in ["DBDS", "SBDS", "GDS"]:
            breaker_kit, quantity = padmount_breaker_pull(one_line, amperage, is_100_percent_rated)

    if erms == True:
        erms_kit = ERMS(voltage_code, one_line, enclosure_code)
    elif erms == False:
        erms_kit = []
    
    print(f"PRM parts are {prm_list}")
    print(f"Adder Parts List: {adder_parts}")
    rotary_parts = rotary_adders(part_list)
    compressions = internal_wiring(amperage, one_line, voltage_code, enclosure_code, is_4P)
    tmts_kit = full_tmts(amperage, voltage_code, part_list)
    full_list = chain(adder_parts, prm_list, breaker_kit, rotary_parts, erms_kit, tmts_kit, temp_gen_kit)
    full_list = list(full_list)
    full_list.append(compressions)
    print(f'Orginial: {full_list}')
    full_list = [entry for entry in full_list if entry[1] != 0]
    final_list = flatten_duplicates(full_list)
    final_list.sort()
    print(f'New: {final_list}')
    epicor_list, quantity_list = split_list(final_list)

## Excel Dumping
if purpose_validity == True and modifier_validity == True and checkbox_validity == True:
    excel_sheet.column_dimensions['A'].width = 25
    excel_sheet.cell(column=1, row=1, value=part_number)
    excel_sheet.cell(column=1, row=2, value="Part")
    excel_sheet.cell(column=2, row=2, value="Quantity")

    for i, value in enumerate(epicor_list):
        excel_sheet.cell(column=1, row=i+3, value=value)
    for i, value in enumerate(quantity_list):
        excel_sheet.cell(column=2, row=i+3, value=value)
    excel_wb.save(filename='BOMCompare.xlsx')
    os.startfile('BOMCompare.xlsx')
    print("Excel Sheet populated")
else:
    pass
